#!/usr/bin/env python3
"""
Export ui-manifest.json to XLSX or CSV.

Usage:
    python3 export.py --manifest data/output/ui-manifest.json --format xlsx
    python3 export.py --manifest data/output/ui-manifest.json --format csv
    python3 export.py --manifest data/output/ui-manifest.json --format xlsx \
        --verification data/output/verification.json

Output file is written next to the manifest with the job name.
"""

import argparse
import csv
import json
import sys
from pathlib import Path


def load_manifest(filepath: Path) -> dict:
    """Load and validate ui-manifest.json."""
    if not filepath.exists():
        print(f"Manifest not found: {filepath}")
        sys.exit(1)

    with open(filepath, "r", encoding="utf-8") as f:
        manifest = json.load(f)

    if manifest.get("schema_version") != "1.0":
        print("Warning: unexpected schema version")

    return manifest


def load_verification(filepath: Path) -> dict:
    """Load verification.json if it exists."""
    if not filepath or not filepath.exists():
        return {}

    with open(filepath, "r", encoding="utf-8") as f:
        data = json.load(f)

    return data.get("cells", {})


def get_cell_value(row: dict, col_id: str) -> str:
    """Extract display value from a row's cell."""
    cell = row.get("cells", {}).get(col_id, {})
    return cell.get("display", cell.get("value", "")) or ""


def get_verification_status(
    verifications: dict, doc_id: str, col_id: str
) -> str:
    """Get verification status for a cell."""
    key = f"{doc_id}:{col_id}"
    entry = verifications.get(key, {})
    return entry.get("status", "")


def export_csv(manifest: dict, verifications: dict, output_path: Path):
    """Export manifest to CSV."""
    columns = manifest["columns"]
    rows = manifest["rows"]

    headers = ["Document"]
    for col in columns:
        headers.append(col["label"])
        headers.append(f"{col['label']} (Status)")

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)

        for row in rows:
            csv_row = [row["_document"]]
            for col in columns:
                csv_row.append(get_cell_value(row, col["id"]))
                status = get_verification_status(
                    verifications, row["_id"], col["id"]
                )
                csv_row.append(status)
            writer.writerow(csv_row)

    print(f"CSV exported: {output_path} ({len(rows)} rows)")


def export_xlsx(manifest: dict, verifications: dict, output_path: Path):
    """Export manifest to XLSX with colour-coded verification."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("Install openpyxl: pip install openpyxl")
        sys.exit(1)

    wb = Workbook()
    ws = wb.active
    ws.title = manifest["job"]["name"][:31]  # Excel tab limit

    columns = manifest["columns"]
    rows = manifest["rows"]

    # Styles
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill("solid", fgColor="D9E2F3")
    verified_fill = PatternFill("solid", fgColor="E2EFDA")
    flagged_fill = PatternFill("solid", fgColor="FCE4EC")
    wrap = Alignment(wrap_text=True, vertical="top")

    # Headers
    ws.cell(row=1, column=1, value="Document").font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    ws.column_dimensions["A"].width = 35

    for i, col in enumerate(columns, start=2):
        cell = ws.cell(row=1, column=i, value=col["label"])
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[cell.column_letter].width = 25

    # Data rows
    for row_idx, row in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=row["_document"])

        for col_idx, col in enumerate(columns, start=2):
            value = get_cell_value(row, col["id"])
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = wrap

            # Apply verification colour
            status = get_verification_status(
                verifications, row["_id"], col["id"]
            )
            if status == "verified":
                cell.fill = verified_fill
            elif status == "flagged":
                cell.fill = flagged_fill

    # Summary sheet
    ws_summary = wb.create_sheet("Summary")
    summary = manifest["summary"]
    summary_data = [
        ("Job Name", manifest["job"]["name"]),
        ("Created", manifest["job"]["created_at"]),
        ("Documents", manifest["job"]["document_count"]),
        ("Columns", manifest["job"]["column_count"]),
        ("Total Cells", summary["total_cells"]),
        ("Complete", summary["complete"]),
        ("Failed", summary["failed"]),
        ("Verified", summary["verified"]),
        ("Flagged", summary["flagged"]),
    ]
    for i, (label, val) in enumerate(summary_data, start=1):
        ws_summary.cell(row=i, column=1, value=label).font = header_font
        ws_summary.cell(row=i, column=2, value=val)

    wb.save(output_path)
    print(f"XLSX exported: {output_path} ({len(rows)} rows)")


def main():
    parser = argparse.ArgumentParser(description="Export review results")
    parser.add_argument(
        "--manifest", required=True, help="Path to ui-manifest.json"
    )
    parser.add_argument(
        "--format",
        required=True,
        choices=["csv", "xlsx"],
        help="Export format",
    )
    parser.add_argument(
        "--verification", help="Path to verification.json (optional)"
    )
    parser.add_argument("--output", help="Output file path (auto-generated)")
    args = parser.parse_args()

    manifest = load_manifest(Path(args.manifest))
    verifications = {}
    if args.verification:
        verifications = load_verification(Path(args.verification))

    # Generate output path
    if args.output:
        output_path = Path(args.output)
    else:
        job_name = manifest["job"]["name"]
        safe_name = job_name.replace(" ", "_").replace("/", "_").lower()
        ext = ".xlsx" if args.format == "xlsx" else ".csv"
        output_path = Path(args.manifest).parent / f"{safe_name}_export{ext}"

    if args.format == "csv":
        export_csv(manifest, verifications, output_path)
    else:
        export_xlsx(manifest, verifications, output_path)


if __name__ == "__main__":
    main()
